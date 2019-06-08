import click
import openpyxl
import django
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE',"classproject.settings")
django.setup()

from onlineapp.models import College,Student,MockTest1

# def get_data(sheet):
#      rangeSelected = []
#      for row in sheet.iter_rows():
#          rowSelected = []
#          for cell in row:
#              rowSelected.append(cell.value)
#          rangeSelected.append(rowSelected)
#      return rangeSelected
#
# @click.group()
# def commands():
#      pass
#
# @commands.command(short_help="create a new database")
# @click.argument("students_excel")
# @click.argument("mock_res_excel")
# def importdata(students_excel,mock_res_excel):
        #read data from student.xlsx
        # workbook_1 = openpyxl.load_workbook(students_excel)
        # sheet_obj_1 = workbook_1.get_sheet_by_name("Colleges")
        # college_data = get_data(sheet_obj_1)[1:]
        # for i in range(len(college_data)):
        #      obj = College(name=college_data[i][0],acronym=college_data[i][1],location = college_data[i][2],contact=college_data[i][3])
        #      obj.save()

        #read student data from student.xlsx
         # workbook_1 = openpyxl.load_workbook(students_excel)
         # sheet_obj_2 = workbook_1.get_sheet_by_name("Deletions")
         # student_data = get_data(sheet_obj_2)[1:]
         # print(student_data)
         # for i in range(len(student_data)):
         #      try:
         #          colleges = College.objects.get(acronym=student_data[i][1])
         #          college_id = int(colleges.id)
         #          obj = Student(name=student_data[i][0],college = colleges,email=student_data[i][2],db_folder=student_data[i][3],dropped_out=1)
         #          obj.save()
         #      except College.DoesNotExist:
         #          print("Error")

        # read mock marks
        # workbook_2 = openpyxl.load_workbook(mock_res_excel)
        # sheet_obj_3 = workbook_2.get_sheet_by_name("Test resul")
        # results_data = get_data(sheet_obj_3)[1:]
        # for i in range(len(results_data)):
        #     temp = results_data[i][0].split("_")[-2]
        #     student = Student.objects.get(db_folder = temp)
        #     obj = MockTest1(student=student,problem1=results_data[i][1],problem2=results_data[i][2],problem3=results_data[i][3],problem4=results_data[i][4],total=results_data[i][5])
        #     obj.save()


#commands()

# c = College.objects.all()
# print(c)

from django.db import connection

def lq():
    return connection.queries[:-1]
lq()